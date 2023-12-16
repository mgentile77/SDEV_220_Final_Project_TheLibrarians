"""
All work on the back end completed and reviewed by
Michael Gentile
"""
#Import librayies required to work with database
import pandas as pd
from openpyxl import load_workbook
import os
#Set a Global decimal format within all dataframes created
pd.set_option('display.float_format', '{:.2f}'.format)


class Library: #Set up pathway for all modules used
    
    def __init__(self):
        self.cwd = os.path.abspath(os.path.join(os.getcwd()))
        self.excel_file = os.path.join(self.cwd, "Library.xlsx")
        
            
class Lib_Books(Library):#Build module for working with and in Book database
    
    def __init__(self): #Initiate the book database to a dataframe
        super().__init__()
        self.bfile = pd.read_excel(self.excel_file, header=0, index_col='Book_ID', sheet_name="Books", dtype={'ISBN':str})       

    def search_bkID(self, bookid): #Query dataframe for a specific book by the book index
        try:
            list = self.bfile.query('Book_ID == [@bookid]')
            if list.empty == True:
                raise Exception("Data is empty")
            return list.to_string(index=True, max_colwidth=45, show_dimensions=True,
                                                justify='center', index_names=True )
        except Exception as e:
            return "Your book was not found.\n\
Please alter your search and try again."
    
    
    def search_book(self, title = "", authfn = "", authln = ""):#Query dataframe for a list of possible books 
        try:
            list = self.bfile.query('Title.str.contains(@title) &\
                Author.str.contains(@authfn)& Author.str.contains(@authln)')
            if list.empty == True:
                raise Exception("Data is empty")
            return list.to_string(index=True, max_colwidth=45, show_dimensions=True,
                                                justify='center', index_names=True )
        except Exception as e:
            return "Your book was not found.\n\
Please alter your search and try again."
    
    def add_book(self, ISBN, title, author, published, publisher): #Add book to database
        new_book_dict = pd.DataFrame({"Book_ID" : [self.bfile.index.max() +3], #Builds temp dataframe
                                          "ISBN" : [ISBN],
                                          "Title" : [title],
                                          "Author" : [author],
                                          "Year-of-Publication" : [int(published)],
                                          "Publisher" : [publisher],
                                          "Availabilty" : ["Available"]})
        #Formats dataframe before adding to database
        new_book_dict = new_book_dict.set_index("Book_ID")
        new_book_dict['ISBN'] = new_book_dict['ISBN'].str.pad(width=10, fillchar='0')
        new_book_dict['Title'] = new_book_dict['Title'].str.upper()
        new_book_dict['Author'] = new_book_dict['Author'].str.upper()
        new_book_dict['Publisher'] = new_book_dict['Publisher'].str.upper()
        new_book_dict['Availabilty'] = new_book_dict['Availabilty'].str.upper()
        #writes to database
        with pd.ExcelWriter(self.excel_file, mode = 'a', if_sheet_exists='overlay') as writer:
            new_book_dict.to_excel(writer, sheet_name='Books', startrow=writer.sheets['Books'].max_row, header=False)
        return f"\nNew Book ID is : {new_book_dict.index.max()}\n"
            
    def drop_book(self, bookid):#Removes book from database
        try:
            self.bfile.loc[bookid]
            __file = self.bfile.drop(bookid)
            #Rewrites database without book
            with pd.ExcelWriter(self.excel_file, mode = 'a', if_sheet_exists='replace') as writer:
                __file.to_excel(writer, sheet_name='Books', header=True)
            return "Book successfully dropped."
        except KeyError:
            return "The book ID entered does not exist.\n\
Please check the number and try again"

    def get_apa_format(self, book_id):#Builds a string in APA format of a given indexed book
        try:
            list = self.bfile.query('Book_ID == [@book_id]')
            if list.empty == True:
                raise KeyError()       
            _author = self.bfile.loc[book_id, 'Author']
            if ' ' in _author:
                    _authLn = _author.split(' ', -1)[-1]
                    _authFn = _author.split(' ', 1)[0]
            else:
                    _authLn = _author
                    _authFn = ''
            _year = int(self.bfile.loc[book_id, 'Year-Of-Publication'])
            _title = self.bfile.loc[book_id, 'Title']
            _publisher = self.bfile.loc[book_id, 'Publisher']
            return f"""{_authLn}, {_authFn[0]}. ({_year}). {_title}. \n\t{_publisher}."""
        except KeyError:
            return  "Your book was not found.\n\
Please alter your search and try again."
        except Exception as e:
            return str(e)

class Lib_Members(Library): #Builds module for working in and with Member database
       
    def __init__(self): #Initiate the member database to a dataframe
        super().__init__()
        self.mfile = pd.read_excel(self.excel_file, index_col='Member_ID',sheet_name="Members")

    def search_memID(self, memid):#Query dataframe for a specific member by the member index
        try:
            list = self.mfile.query('Member_ID == @memid')
            if list.empty == True:
                raise Exception("Data is empty")
            return list.to_string(index=True, max_colwidth=45, show_dimensions=True,
                                                justify='center', index_names=True )
        except Exception as e:
            return "Member was not found.\n\
Please alter your search and try again"
    
    def search_mem(self, memfn = "", memln = "", memphone = ""):#Query dataframe for a list of possible members
        try:
            list = self.mfile.query('Fname.str.startswith(@memfn) &\
                Lname.str.startswith(@memln)& Phone.str.endswith(@memphone)')
            if list.empty == True:
                raise Exception("Data is empty")
            return list.to_string(index=True, max_colwidth=45, show_dimensions=True,
                                                justify='center', index_names=True )
        except Exception as e:
            return "Member was not found.\n\
Please alter your search and try again"
    
    def add_mem(self, fn, ln, stadd, city, state, zip, email, phone):#Add member to database
        new_mem_df = pd.DataFrame({"Member_ID" : [self.mfile.index.max()+3],#Builds temp dataframe
                                          "Fname" : [fn],"Lname" : [ln],
                                          "Street_Add" : [stadd],"City" : [city],
                                          "State": [state],"Zip" : [zip],
                                          "Email":[email], "Phone": [phone]})
        #Formats dataframe before adding to database
        new_mem_df = new_mem_df.set_index("Member_ID")
        new_mem_df['Fname'] = new_mem_df['Fname'].str.title()
        new_mem_df['Lname'] = new_mem_df['Lname'].str.title()
        new_mem_df['Street_Add'] = new_mem_df['Street_Add'].str.title()
        new_mem_df['City'] = new_mem_df['City'].str.title()
        new_mem_df['State'] = new_mem_df['State'].str.upper()        
        new_mem_df['Email'] = new_mem_df['Email'].str.lower()
        #writes to database
        with pd.ExcelWriter(self.excel_file, mode = 'a', if_sheet_exists='overlay') as writer:
            new_mem_df.to_excel(writer, sheet_name='Members', startrow=writer.sheets['Members'].max_row, header=False)
        return f"\nNew Member ID is : {new_mem_df.index.max()}\n"
            
    def drop_mem(self, memid):#Removes book from database
        try:
            self.mfile.loc[memid]
            __file = self.mfile.drop(memid)
            #Rewrites database without book
            with pd.ExcelWriter(self.excel_file, mode = 'a', if_sheet_exists='replace') as writer:
                __file.to_excel(writer, sheet_name='Members', header=True)
            return "Member(s) successfully dropped"
        except KeyError:
            return "The Member ID entered does not exist.\n\
Please check the number and try again."
            
            
class Lib_Fines(Library):#Builds module for working in and with Fines database
    def __init__(self):
        super().__init__()#Initiate the member database to a dataframe
        self.fine_rate_per_day = 0.05#sets fine rate
        self.finefile = pd.read_excel(self.excel_file, header=0, sheet_name="Fines", index_col=[0])
        self.borchk = pd.read_excel(self.excel_file, header=0, sheet_name="Books_Out", 
                                index_col=[0], parse_dates=['Date_Borrowed', 'Date_Due'])
        self.borchk = self.borchk.reset_index()#Resets dataframe index for the specific functions for working with database
        
    def assess_fines(self):#Automates fine assessment
        fines_per_member = {}
        for row in self.borchk.index:#Adds new members to list with late items to assess fines
            member_id = self.borchk.loc[row, 'Member_ID']
            if self.borchk.loc[row,'Late'] == 'Late':
                fine = self.fine_amount(row)
                fines_per_member[member_id] = fines_per_member.get(member_id, 0) + fine
        for k, v in fines_per_member.items():#Adds fines to members already late and manages balances
            if k in self.finefile.index.get_level_values('Member_ID'):
                self.finefile.loc[k, 'Fine_Accrued'] = v
                self.finefile.loc[k, 'Balance'
                                  ] = self.finefile.loc[k, 'Fine_Accrued'
                                                        ] - self.finefile.loc[k, 'Payments_Received']
            else:#Initiates columns for new members added
                self.finefile.loc[k] = {'Fine_Accrued': v, 'Payments_Received': 0.00, 'Balance': v}
        return self.finefile
    
    def clear_empty_balance(self):#Automates removal of members from fine database when balance is 0
        _df = self.assess_fines()
        for row in _df.index:
            if _df.loc[row,'Balance'] == 0:
                _df.drop(row, inplace=True)
        #Rewrites database
        with pd.ExcelWriter(self.excel_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            _df.to_excel(writer, sheet_name='Fines', index=True)

    def fine_amount(self, row):#Determines how to determine if a member should be fined
        current_date = pd.Timestamp.now().floor('D')
        due_date = self.borchk.loc[row, 'Date_Due']
        total_days = (current_date - due_date).days
        return total_days * self.fine_rate_per_day
    
    def update_fine_record(self,member_id, amount):#Allows member to pay off balance
        try:
            _fines = self.finefile.loc[member_id, 'Fine_Accrued']
            if amount > _fines:
                self.finefile.loc[member_id, 'Payments_Received'] = _fines
                self.finefile.loc[member_id, 'Balance'] = 0.00
                return f"Change: ${abs(amount - _fines):.2f}"
            self.finefile.loc[member_id, 'Payments_Received'] += amount
            self.finefile.loc[member_id, 'Balance'] = _fines - self.finefile.loc[member_id,'Payments_Received']
            _balance = self.finefile.loc[member_id, 'Balance']
            return f"Balance Remaining: ${_balance:.2f}"
        except KeyError:
            return "The Member Id was not found to have any fines.\nPlease check the ID and try again."

class Report(Lib_Books, Lib_Members, Lib_Fines):#Builds module for viewing Reports
    def __init__(self): #Iniates all databases required for reports
        super().__init__()
        self.identify_late()
        self.clear_empty_balance()
        self.borfile = pd.read_excel(self.excel_file, header=0, index_col=[0,1], sheet_name="Books_Out",
                                     parse_dates = ['Date_Borrowed', 'Date_Due'])
       
    def identify_late(self): #Updates book return statue
        self.borfile = pd.read_excel(self.excel_file, header=0, index_col=[0,1],
                                sheet_name="Books_Out", parse_dates = ['Date_Borrowed', 'Date_Due'])
        self.borfile.loc[self.borfile['Date_Due'] < pd.Timestamp.now(), 'Late'] = 'Late'
        self.borfile.loc[self.borfile['Date_Due'] > pd.Timestamp.now(), 'Late'] = ''
        self.borfile = self.borfile.sort_values(['Member_ID', 'Date_Due'])
        #Rewrites database aafter update
        with pd.ExcelWriter(self.excel_file, mode = 'a', date_format='MM/DD/YYYY', if_sheet_exists='replace') as writer:
            self.borfile.to_excel(writer, sheet_name='Books_Out', header=True)
    
    def bks_out_emails(self): #Prepares report for books that are out and members actively using library
        #Builds temp dataframe
        __all_merge = self.borfile.merge(self.mfile[['Fname', 'Lname', 'Email']], right_index= True, 
                                            left_on= 'Member_ID'
                                            ).merge(self.bfile[['Title', 'Author', 'ISBN']], right_index= True, 
                                                    left_on= 'Book_ID').reset_index(level='Book_ID')
        #Formats dataframe
        __all_merge['Name']= (__all_merge['Fname']+" "+__all_merge['Lname'])
        __all_merge['Author'] = __all_merge['Author'].str.split(n=-1).str[-1]
        __all_merge = __all_merge[['Name', 'Email', 'ISBN','Title', 'Author', 'Date_Due','Late']
                                    ].set_index(['Name', 'Email', 'ISBN'], append=True)
        __all_merge= __all_merge.dropna(axis=0, how='any')
        #Converts dataframe to string for viewing
        __all_merge = __all_merge.to_string(index=True, max_colwidth=45, show_dimensions=True,
                                            justify='left', index_names=True, 
                                            formatters={'Title': (lambda x: '{:<45}'.format(x)),'Author': (lambda x: '{:<15}'.format(x))})
        return f'{__all_merge}\n\n'

    def bks_late(self):#Prepares report for books that are out and members actively using library
        #Builds temp dataframe    
        __late_merge = self.borfile.merge(self.mfile[['Fname', 'Lname', 'Phone', 'Email']], right_index=True, 
                                            left_on='Member_ID'
                                            ).merge(self.bfile[['Title', 'Author']], right_index=True, 
                                                    left_on= 'Book_ID'
                                                    ).merge(self.finefile[['Balance']], left_on='Member_ID', right_index = True)
        #Formats dataframe
        __late_merge[['Days_Late']] = pd.Timestamp.today().floor('D') - __late_merge[['Date_Due']]
        __late_merge['Name']= (__late_merge['Fname']+" "+__late_merge['Lname'])
        __late_merge['Author'] = __late_merge['Author'].str.split(n=-1).str[-1]
        __late_merge = __late_merge[['Name', 'Phone', 'Email', 'Title', 'Author','Days_Late', 'Late','Balance']]
        __late_merge= __late_merge.set_index(['Name', 'Phone', 'Email','Balance', 'Title']
                                                ).groupby(['Days_Late']).filter(lambda x: (x['Late'] == 'Late').any())
        __late_merge=__late_merge.drop(columns='Late')
        __late_merge= __late_merge.dropna(axis=0, how='any')
        #Converts dataframe to string for viewing
        __late_merge = __late_merge.to_string(index=True, max_colwidth=45, show_dimensions=True,
                                            justify='left', index_names=True, 
                                            formatters={'Title': (lambda x: '{:<45}'.format(x)),'Author': (lambda x: '{:<15}'.format(x))})
        return f'{__late_merge}\n\n'
    


class Lib_Borrow(Report, Lib_Fines):#Builds module for working in and with Borrowed Book database
    
    def __init__(self):#Iniates from parent classes
        super().__init__()
    
    def bor_bk(self, memid, bkid):#Allows member to borrow book
        try:#checks to see if book is available
            self.bfile.loc[bkid]
            self.mfile.loc[memid]
            _cur_date = pd.Timestamp.now().floor(freq='D')
            _due = _cur_date + pd.DateOffset(days=14)
            if bkid in self.borfile.index.get_level_values('Book_ID'):
                return f"""We apologize for the inconvience, but your book is not available at this time.
    Book ID: {bkid}"""
            else:#Builds temp dataframe for borrowing a book
                _bor_df = pd.DataFrame({'Member_ID':[memid],
                            'Book_ID':[bkid],
                            'Date_Borrowed':[_cur_date],
                            'Date_Due':[_due],
                            'Late':['']}).set_index(['Member_ID', 'Book_ID'])
                _new_df = pd.concat([self.borfile,_bor_df], names=['Member_ID', 'Book_ID'])#Adds additional books to datframe of borrowed books
                #Rewrites books to database
                with pd.ExcelWriter(self.excel_file, mode = 'a', if_sheet_exists='replace') as writer:
                    _new_df.to_excel(writer, sheet_name='Books_Out', header=True)
                self.bfile.loc[bkid,'Availability'] = 'OUT'#Marks availability in Book Database
                #Rewrites book database
                with pd.ExcelWriter(self.excel_file, mode = 'a', if_sheet_exists='replace') as writer:
                    self.bfile.to_excel(writer, sheet_name='Books', header=True)
                return "Book added to receipt."
        except KeyError:
            return "Either the Book ID or Member ID is incorrect.\nPlease check the numbers and try again."
                
    def ret_bk(self, bkid):#Allows member to return book
        if bkid in self.borfile.index.get_level_values('Book_ID'):#Removes book from borrow database
            self.borfile.drop(bkid, axis=0, level= 1, inplace=True)
            #Rewrites borrow database
            with pd.ExcelWriter(self.excel_file, mode = 'a', if_sheet_exists='replace') as writer:
                self.borfile.to_excel(writer, sheet_name='Books_Out', header=True)
            self.bfile.loc[bkid,'Availability'] = 'AVAILABLE' #Marks availability in book database
            #Rewrites book databse
            with pd.ExcelWriter(self.excel_file, mode = 'a', if_sheet_exists='replace') as writer:
                self.bfile.to_excel(writer, sheet_name='Books', header=True)
            return 'Book has been returned'
        return f"""That book has not been borrowed.  Please check catalog for availability.
Book ID: {bkid}"""

    def member_receipt(self, memid):#Builds receipt for member, also used for quick view of member records
        #Builds receipt and collects balance for member if late fees are owed
        if (memid in self.borfile.index.get_level_values('Member_ID')) and (memid in self.finefile.index.get_level_values('Member_ID')):
            #Build temp dataframe
            __receipt = self.borfile.merge(self.mfile[['Fname', 'Lname']], right_index=True, 
                                            left_on='Member_ID'
                                            ).merge(self.finefile[['Balance']], left_on='Member_ID', right_index = True
                                                            ).merge(self.bfile[['Title', 'Author']],
                                                                    right_index=True,left_on= 'Book_ID'
                                                                    ).reset_index(level='Book_ID')
            #Format dataframe
            __receipt['Name']= (__receipt['Fname']+" "+__receipt['Lname'])
            __receipt = __receipt[['Name', 'Title', 'Author', 'Date_Due', 'Balance']]
            __receipt= __receipt.loc[[memid]].set_index(['Name', 'Balance', 'Title'], append=True).dropna()
            _rec_count = __receipt.shape[0]
            #Convert to string output for viewing
            _rec_count =  __receipt.to_string(index=True, max_colwidth=45, show_dimensions=True,
                                            justify='center', index_names=True )
            return f"{__receipt}\n\nYou have borrowed {_rec_count} books.\nPlease remember your due dates.\n\n\n"
        #Builds receipt for member if no late fees are owed
        elif memid in self.borfile.index.get_level_values('Member_ID'):
            #Build temp dataframe
            __receipt = self.borfile.merge(self.mfile[['Fname', 'Lname']], right_index=True, 
                                            left_on='Member_ID'
                                            ).merge(self.bfile[['Title', 'Author']], right_index=True, 
                                                    left_on= 'Book_ID'
                                                    ).reset_index(level='Book_ID')
            #Format dataframe
            __receipt['Name']= (__receipt['Fname']+" "+__receipt['Lname'])
            __receipt = __receipt[['Name', 'Title', 'Author', 'Date_Due']]
            __receipt= __receipt.loc[[memid]].set_index(['Name', 'Title'], append=True).dropna()
            _rec_count = __receipt.shape[0]
            #Convert to string output for viewing
            _rec_count =  __receipt.to_string(index=True, max_colwidth=45, show_dimensions=True,
                                            justify='center', index_names=True )
            return f"{__receipt}\n\nYou have borrowed {_rec_count} books.\nPlease remember your due dates.\n\n\n"

        else:
            return "Member does not have any books out."